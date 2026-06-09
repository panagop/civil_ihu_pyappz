"""Build the September 2026 ΔΙΠΑΕ exam schedule with CP-SAT.

Hard constraints (restraints_2026-09.md):
  - Window 01..23 Sep 2026, weekdays only; slots 09/12/15/18.
  - Courses come from all_courses.json; instructors from the ΔΙΠΑΕ sheet.
  - Courses without an instructor (blank / ΔΕΠ) are NOT scheduled.
  - ΚΥ / ΧΥ courses are examined alone (no other exam at the same day+time).
  - Electives (ΔΥ ΓΥ ΣΥ ΥΥ ΔΕ ΓΕ ΣΕ ΥΕ) may share a slot only if they do not
    share a direction (1st letter of the type: Δ Γ Σ Υ) within the same semester.
  - Same instructor never in two places at once.
  - Per study-stream spacing: >= 1 free day between exams a single student sits
    (weekends count as free days). A stream = the ΚΥ/ΧΥ of a semester, plus, for
    semesters 7-9, the electives of each direction.
  - Δανιήλ: only Tue/Fri.  Καζαντζή: only 02-04, 11, 21-23 Sep.
  - Βλαχονάσιου/Φωτοπούλου/Δανιήλ: not 18:00.  Βοζίκης: 12:00 or 15:00.

Soft (objective, "if possible"):
  - Grouping instructors get 2 courses on the same day.
  - Λιαλιαμπής on the same day as Βλαχονάσιου, in an adjacent slot.
  - Even distribution within each stream (maximise the minimum gap).
"""
import sys, io, json, os
from datetime import date, timedelta
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import pandas as pd
from ortools.sat.python import cp_model

FILE = "files/exams/exams-2026-09.xlsm"
START, END = date(2026, 9, 1), date(2026, 9, 23)
SLOTS = ["09:00", "12:00", "15:00", "18:00"]          # slot index 0..3
ELECTIVE_TYPES = {"ΔΥ", "ΓΥ", "ΣΥ", "ΥΥ", "ΔΕ", "ΓΕ", "ΣΕ", "ΥΕ"}
GROUP_INSTR = {"Αυγέρης", "Κοκκαλά", "Κόκκινος", "Μπακάλης",
               "Παπαϊωάννου", "Σαπίδης", "Τσιαράπας", "Φαναραδέλλη",
               "Βοζίκης", "Βλαχονάσιου", "Δανιήλ", "Μιχαηλίδης"}

# ---- valid days as ordinals (days since START), weekdays only ----------------
days = [d for d in range((END - START).days + 1)
        if (START + timedelta(d)).weekday() < 5]
def ordinal(d): return (d - START).days
def to_date(o): return START + timedelta(o)
KAZANTZI_DAYS = [ordinal(date(2026, 9, x)) for x in (2, 3, 4, 11, 21, 22, 23)]
TUE_FRI = [o for o in days if to_date(o).weekday() in (1, 4)]
NO_1800 = {"Βλαχονάσιου", "Φωτοπούλου", "Δανιήλ"}

# ---- load courses to schedule -----------------------------------------------
with open("files/exams/input/all_courses.json", encoding="utf-8") as fh:
    by_code = {c["code"]: c for c in json.load(fh)}
sheet = pd.read_excel(FILE, sheet_name="ΔΙΠΑΕ")

def real_instr(v):
    if pd.isna(v): return None
    s = str(v).strip()
    return None if s == "" or s.upper() == "ΔΕΠ" else s

C = []  # list of dicts: id, name, sem, instr, types, dirs, alone
for _, r in sheet.iterrows():
    cid = r["course_id"]
    if pd.isna(cid):
        continue
    cid = str(cid).strip()
    instr = real_instr(r["instructor"])
    info = by_code.get(cid)
    if instr is None or info is None:
        continue
    types = info["types"]
    dirs = {t[0] for t in types if t in ELECTIVE_TYPES}
    alone = any(t in ("ΚΥ", "ΧΥ", "ΠΥ") for t in types)
    C.append(dict(id=cid, name=str(r["course_name"]), sem=int(info["semester"]),
                  instr=instr, types=types, dirs=dirs, alone=alone))
N = len(C)
print(f"scheduling {N} courses; alone={sum(c['alone'] for c in C)}, "
      f"electives={sum(not c['alone'] for c in C)}")

# ---- build streams -----------------------------------------------------------
streams = []
for sem in range(1, 7):
    s = [i for i, c in enumerate(C) if c["sem"] == sem]
    if s: streams.append(s)
for sem in (7, 8, 9):
    for d in ("Δ", "Γ", "Σ", "Υ"):
        s = [i for i, c in enumerate(C)
             if c["sem"] == sem and (c["alone"] or d in c["dirs"])]
        if len(s) > 1:
            streams.append(s)
print("streams:", len(streams), "max stream size:", max(len(s) for s in streams))

# ---- model -------------------------------------------------------------------
m = cp_model.CpModel()
day = []
slot = []
cell = []
for i, c in enumerate(C):
    dom = days
    if c["instr"] == "Δανιήλ":
        dom = TUE_FRI
    elif c["instr"] == "Καζαντζή":
        dom = KAZANTZI_DAYS
    elif c["instr"] == "Λιαλιαμπής":
        dom = [o for o in days if to_date(o) < date(2026, 9, 15)]
    dv = m.NewIntVarFromDomain(cp_model.Domain.FromValues(dom), f"day_{i}")
    if c["instr"] == "Βοζίκης":
        sv = m.NewIntVarFromDomain(cp_model.Domain.FromValues([1, 2]), f"slot_{i}")
    elif c["instr"] in NO_1800:
        sv = m.NewIntVarFromDomain(cp_model.Domain.FromValues([0, 1, 2]), f"slot_{i}")
    else:
        sv = m.NewIntVar(0, 3, f"slot_{i}")
    cv = m.NewIntVar(0, 4 * max(days) + 3, f"cell_{i}")
    m.Add(cv == 4 * dv + sv)
    day.append(dv); slot.append(sv); cell.append(cv)

alone = [i for i, c in enumerate(C) if c["alone"]]

# (B) alone courses occupy a globally unique cell
for a in alone:
    for x in range(N):
        if x != a:
            m.Add(cell[a] != cell[x])

# same instructor never overlaps
from collections import defaultdict
by_instr = defaultdict(list)
for i, c in enumerate(C):
    by_instr[c["instr"]].append(i)
for ids in by_instr.values():
    for a in range(len(ids)):
        for b in range(a + 1, len(ids)):
            m.Add(cell[ids[a]] != cell[ids[b]])

# electives same semester sharing a direction -> different cell
for i in range(N):
    for j in range(i + 1, N):
        ci, cj = C[i], C[j]
        if ci["alone"] or cj["alone"]:
            continue
        if ci["sem"] == cj["sem"] and (ci["dirs"] & cj["dirs"]):
            m.Add(cell[i] != cell[j])

# spacing within streams (hard: >= 1 free day between same-stream exams)
for s in streams:
    for a in range(len(s)):
        for b in range(a + 1, len(s)):
            diff = m.NewIntVar(-max(days), max(days), "")
            m.Add(diff == day[s[a]] - day[s[b]])
            ad = m.NewIntVar(0, max(days), "")
            m.AddAbsEquality(ad, diff)
            m.Add(ad >= 2)

# Γαλάνης: his courses fall in two consecutive-day periods, period A spanning
# up to 4 calendar days (day-diff <= 3) and period B up to 3 (day-diff <= 2),
# with A entirely before B so they are two distinct blocks.
gal = by_instr.get("Γαλάνης", [])
if len(gal) > 1:
    inA = {i: m.NewBoolVar(f"galA_{i}") for i in gal}
    for a in range(len(gal)):
        for b in range(a + 1, len(gal)):
            i, j = gal[a], gal[b]
            m.Add(day[i] - day[j] <= 3).OnlyEnforceIf([inA[i], inA[j]])
            m.Add(day[j] - day[i] <= 3).OnlyEnforceIf([inA[i], inA[j]])
            m.Add(day[i] - day[j] <= 2).OnlyEnforceIf([inA[i].Not(), inA[j].Not()])
            m.Add(day[j] - day[i] <= 2).OnlyEnforceIf([inA[i].Not(), inA[j].Not()])
            m.Add(day[i] < day[j]).OnlyEnforceIf([inA[i], inA[j].Not()])
            m.Add(day[j] < day[i]).OnlyEnforceIf([inA[i].Not(), inA[j]])
    m.Add(sum(inA[i] for i in gal) >= 1)
    m.Add(sum(inA[i] for i in gal) <= len(gal) - 1)

# even distribution (soft): flatten per-day load + cap parallelism per slot
is_on = {}                      # is_on[i, o] == 1 iff course i is on day o
for i in range(N):
    for o in days:
        b = m.NewBoolVar(f"on_{i}_{o}")
        m.Add(day[i] == o).OnlyEnforceIf(b)
        m.Add(day[i] != o).OnlyEnforceIf(b.Not())
        is_on[i, o] = b
loads = []
for o in days:
    ld = m.NewIntVar(0, N, f"load_{o}")
    m.Add(ld == sum(is_on[i, o] for i in range(N)))
    loads.append(ld)
maxload = m.NewIntVar(0, N, "maxload")
m.AddMaxEquality(maxload, loads)
# hard: grouping instructors get at most 2 courses on the same day.
# STRICT_PAIR=1  -> every used day must hold exactly 2 (no singletons at all).
# STRICT_PAIR=2  -> mandatory pairing with at most ONE leftover single per
#                   instructor (the only feasible form when a count is odd).
STRICT_PAIR = os.environ.get("STRICT_PAIR", "0")
SINGLETONS = []                          # mode 3: per-day unpaired indicators
for instr in GROUP_INSTR:
    ids = by_instr.get(instr, [])
    if not ids:
        continue
    if STRICT_PAIR == "1":
        for o in days:
            b = m.NewBoolVar(f"pair_{instr}_{o}")
            m.Add(sum(is_on[i, o] for i in ids) == 2 * b)   # cnt in {0, 2}
    elif STRICT_PAIR in ("2", "3"):
        ones = []
        for o in days:
            one = m.NewBoolVar(f"one_{instr}_{o}")
            two = m.NewBoolVar(f"two_{instr}_{o}")
            m.Add(sum(is_on[i, o] for i in ids) == one + 2 * two)  # cnt in {0,1,2}
            m.Add(one + two <= 1)
            ones.append(one)
        if STRICT_PAIR == "2":
            m.Add(sum(ones) <= 1)        # hard: at most one unpaired course
        SINGLETONS.extend(ones)          # mode 3: minimise these instead
    elif len(ids) > 2:
        for o in days:
            m.Add(sum(is_on[i, o] for i in ids) <= 2)

# hard: Μιχαηλίδης has >= 2 courses in the first week (01-04 Sep)
FIRST_WEEK = [o for o in days if to_date(o) <= date(2026, 9, 4)]
mich = by_instr.get("Μιχαηλίδης", [])
if len(mich) >= 2:
    m.Add(sum(is_on[i, o] for i in mich for o in FIRST_WEEK) >= 2)
# at most 4 exams running in parallel in any one day+time slot
for o in days:
    for sl in range(4):
        par = []
        for i in range(N):
            b = m.NewBoolVar(f"in_{i}_{o}_{sl}")
            m.Add(cell[i] == 4 * o + sl).OnlyEnforceIf(b)
            m.Add(cell[i] != 4 * o + sl).OnlyEnforceIf(b.Not())
            par.append(b)
        m.Add(sum(par) <= 4)
obj = [-maxload]

# soft: grouping instructors -> pairs on the same day
group_terms = []
for instr in GROUP_INSTR:
    ids = by_instr.get(instr, [])
    for a in range(len(ids)):
        for b in range(a + 1, len(ids)):
            eq = m.NewBoolVar(f"grp_{instr}_{a}_{b}")
            m.Add(day[ids[a]] == day[ids[b]]).OnlyEnforceIf(eq)
            m.Add(day[ids[a]] != day[ids[b]]).OnlyEnforceIf(eq.Not())
            group_terms.append(eq)

# soft: Λιαλιαμπής adjacent to Βλαχονάσιου (same day, |slot diff| == 1)
adj_terms = []
for li in by_instr.get("Λιαλιαμπής", []):
    options = []
    for vl in by_instr.get("Βλαχονάσιου", []):
        b = m.NewBoolVar(f"adj_{li}_{vl}")
        sd = m.NewIntVar(-3, 3, "")
        m.Add(sd == slot[li] - slot[vl])
        asd = m.NewIntVar(0, 3, "")
        m.AddAbsEquality(asd, sd)
        m.Add(day[li] == day[vl]).OnlyEnforceIf(b)
        m.Add(asd == 1).OnlyEnforceIf(b)
        options.append(b)
    if options:
        any_adj = m.NewBoolVar(f"adj_any_{li}")
        m.AddBoolOr(options).OnlyEnforceIf(any_adj)
        for o in options:
            m.Add(o == 0).OnlyEnforceIf(any_adj.Not())  # not required, just link
        adj_terms.append(any_adj)

import os
if os.environ.get("NOOBJ") != "1":
    # mode 3: dominate everything else by minimising unpaired singles
    pair_obj = -100000 * sum(SINGLETONS) if SINGLETONS else 0
    m.Maximize(pair_obj + 500 * sum(adj_terms) + 20 * sum(group_terms) + 50 * sum(obj))

solver = cp_model.CpSolver()
solver.parameters.max_time_in_seconds = float(os.environ.get("TLIM", "60"))
solver.parameters.num_search_workers = 8
status = solver.Solve(m)
print("status:", solver.StatusName(status))
if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
    print("NO SOLUTION")
    sys.exit(1)

print(f"adjacency={sum(int(solver.Value(a)) for a in adj_terms)}/{len(adj_terms)} "
      f"grouping={sum(int(solver.Value(g)) for g in group_terms)}/{len(group_terms)}")
if SINGLETONS:
    print(f"unpaired singles (min): {sum(int(solver.Value(s)) for s in SINGLETONS)}")

# ---- emit result -------------------------------------------------------------
out = {}
for i, c in enumerate(C):
    o = solver.Value(day[i]); s = solver.Value(slot[i])
    out[c["id"]] = (to_date(o).isoformat(), SLOTS[s] + ":00")

result = []
for i, c in enumerate(C):
    o = solver.Value(day[i]); s = solver.Value(slot[i])
    result.append((to_date(o), s, c))
result.sort(key=lambda x: (x[0], x[1]))
print("\n date       slot   sem  course   instr           type")
for d, s, c in result:
    star = "*" if c["alone"] else " "
    print(f" {d.isoformat()} {SLOTS[s]} {star} {c['sem']:>2}  {c['id']:7} "
          f"{c['instr']:14} {','.join(c['types'])}")

with open("files/exams/input/_schedule_out.json", "w", encoding="utf-8") as fh:
    json.dump(out, fh, ensure_ascii=False, indent=2)
print("\nwrote files/exams/input/_schedule_out.json")

# ---- write exam_date / start_time back into the ΔΙΠΑΕ sheet of the xlsm -------
import openpyxl
from datetime import datetime
wb = openpyxl.load_workbook(FILE, keep_vba=True)
ws = wb["ΔΙΠΑΕ"]
COL_DATE, COL_TIME = 5, 6                       # E = exam_date, F = start_time
ws.cell(row=1, column=COL_DATE).value = "exam_date"
ws.cell(row=1, column=COL_TIME).value = "start_time"
written = 0
for row in range(2, ws.max_row + 1):
    cid = ws.cell(row=row, column=1).value
    if cid is None:
        continue
    cid = str(cid).strip()
    if cid in out:
        d_iso, t_str = out[cid]
        cell_d = ws.cell(row=row, column=COL_DATE)
        cell_d.value = datetime.fromisoformat(d_iso)
        cell_d.number_format = "yyyy-mm-dd"
        ws.cell(row=row, column=COL_TIME).value = t_str
        written += 1
    else:
        # course not scheduled (no instructor): clear any stale values
        ws.cell(row=row, column=COL_DATE).value = None
        ws.cell(row=row, column=COL_TIME).value = None
wb.save(FILE)
print(f"wrote {written} exam_date/start_time rows into {FILE}")
